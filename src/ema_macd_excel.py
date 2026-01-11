"""Convert weekly CSVs to Excel and add EMA/MACD-H columns."""

from __future__ import annotations

from pathlib import Path

import pandas as pd

from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter

from ema_macd import add_ema_macd_features


def _find_close_column(columns: list[str]) -> str:
    for col in columns:
        if col.lower() == "close":
            return col
    raise ValueError("No 'close' column found in the Excel sheet.")


def _coerce_numeric(series: pd.Series) -> pd.Series:
    if series.dtype == object:
        cleaned = series.astype(str).str.replace(",", "", regex=False)
        multipliers = {"K": 1_000, "M": 1_000_000, "B": 1_000_000_000}
        suffix = cleaned.str[-1].str.upper()
        base = pd.to_numeric(cleaned.str[:-1], errors="coerce")
        numeric = base.mul(suffix.map(multipliers).fillna(1))
        fallback = pd.to_numeric(cleaned, errors="coerce")
        return numeric.where(numeric.notna(), fallback)
    return series.astype(float)


def _coerce_percent(series: pd.Series) -> pd.Series:
    if series.dtype == object:
        return series.astype(str).str.replace("%", "", regex=False).astype(float)
    return series.astype(float)


def _normalize_date(series: pd.Series) -> pd.Series:
    parsed = pd.to_datetime(series, format="%d/%m/%Y", errors="coerce")
    if parsed.isna().all():
        parsed = pd.to_datetime(series, format="%m/%d/%Y", errors="coerce")
    formatted = parsed.dt.strftime("%d/%m/%Y")
    return formatted.where(parsed.notna(), series)


def _trend_from_series(series: pd.Series) -> pd.Series:
    previous = series.shift(1)
    trend = series.gt(previous).map({True: "Up", False: "Down"})
    return trend.where(previous.notna(), pd.NA)


def _extract_etf_name(file_path: Path) -> str:
    stem = file_path.stem
    suffix = " Stock Price History"
    if stem.endswith(suffix):
        stem = stem[: -len(suffix)]
    return stem.replace(" ", "_")


def process_csv(file_path: Path, output_dir: Path) -> None:
    df = pd.read_csv(file_path)
    if "Price" in df.columns and "Close" not in df.columns:
        df = df.rename(columns={"Price": "Close"})

    if "Change %" in df.columns and "Change" not in df.columns:
        df = df.rename(columns={"Change %": "Change"})

    close_col = _find_close_column(list(df.columns))
    df[close_col] = _coerce_numeric(df[close_col])

    # Ensure oldest-first ordering before EMA calculations.
    df = df.iloc[::-1].reset_index(drop=True)
    df_with_features = add_ema_macd_features(df, close_column=close_col)

    date_col = "Date" if "Date" in df_with_features.columns else "date"
    date_series = _normalize_date(df_with_features[date_col])

    volume_col = "Vol." if "Vol." in df_with_features.columns else "Volume"
    if volume_col in df_with_features.columns:
        volume_series = _coerce_numeric(df_with_features[volume_col])
    else:
        volume_series = pd.Series([pd.NA] * len(df_with_features), index=df_with_features.index)

    open_series = _coerce_numeric(df_with_features.get("Open", df_with_features.get("open")))
    high_series = _coerce_numeric(df_with_features.get("High", df_with_features.get("high")))
    low_series = _coerce_numeric(df_with_features.get("Low", df_with_features.get("low")))
    close_series = _coerce_numeric(df_with_features.get("Close", df_with_features.get("close")))

    change_series = (
        _coerce_percent(df_with_features["Change"]) / 100
        if "Change" in df_with_features.columns
        else pd.Series([pd.NA] * len(df_with_features), index=df_with_features.index)
    )

    direction_series = _trend_from_series(close_series)
    ema26_series = df_with_features["ema_26"]
    trend_ema26 = _trend_from_series(ema26_series)
    ema12_series = df_with_features["ema_12"]
    trend_ema12 = _trend_from_series(ema12_series)
    fast_line_series = df_with_features["fast_line"]
    slow_line_series = df_with_features["slow_line"]
    macd_h_series = df_with_features["macd_h"]
    trend_macd_h = _trend_from_series(macd_h_series)

    ordered_columns = [
        date_series,
        volume_series,
        open_series,
        high_series,
        low_series,
        close_series,
        change_series,
        direction_series,
        ema26_series,
        trend_ema26,
        ema12_series,
        trend_ema12,
        fast_line_series,
        slow_line_series,
        macd_h_series,
        trend_macd_h,
    ]
    df_out = pd.concat(ordered_columns, axis=1)
    df_out.columns = [
        "Date",
        "Volume",
        "Open",
        "High",
        "Low",
        "Close",
        "Change",
        "Direction",
        "26-EMA",
        "Trend 26-EMA",
        "12-EMA",
        "Trend 12-EMA",
        "fast line (12-26)",
        "slow line (9-EMA)",
        "MACD-H",
        "Trend MACD-H",
    ]

    etf_name = _extract_etf_name(file_path)
    output_path = output_dir / f"Weekly_{etf_name}.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl", mode="w") as writer:
        df_out.to_excel(writer, sheet_name="Weekly", index=False)
        worksheet = writer.sheets["Weekly"]

        # Apply formatting: font, borders, header alignment, freeze top row, and filter header row.
        font = Font(name="Candara Light")
        header_font = Font(name="Candara Light", bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", start_color="16365C", end_color="16365C")
        no_border = Border()
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = font
                cell.border = no_border

        header_alignment = Alignment(horizontal="left", vertical="center")
        for cell in worksheet[1]:
            cell.alignment = header_alignment
            cell.font = header_font
            cell.fill = header_fill

        worksheet.freeze_panes = "A2"
        last_col = get_column_letter(worksheet.max_column)
        worksheet.auto_filter.ref = f"A1:{last_col}1"

        # Set column widths based on header length for readability.
        for col_idx in range(1, worksheet.max_column + 1):
            header_value = worksheet.cell(row=1, column=col_idx).value
            header_text = str(header_value) if header_value is not None else ""
            worksheet.column_dimensions[get_column_letter(col_idx)].width = max(
                10, len(header_text) + 2
            )

        accounting_columns = {
            "Open",
            "High",
            "Low",
            "Close",
            "26-EMA",
            "12-EMA",
            "fast line (12-26)",
            "slow line (9-EMA)",
            "MACD-H",
        }
        integer_columns = {"Volume"}
        percent_columns = {"Change"}

        for col_idx in range(1, worksheet.max_column + 1):
            header_value = worksheet.cell(row=1, column=col_idx).value
            if header_value in accounting_columns:
                for cell in worksheet.iter_cols(
                    min_col=col_idx, max_col=col_idx, min_row=2
                ):
                    for value_cell in cell:
                        if isinstance(value_cell.value, str):
                            cleaned = value_cell.value.replace(",", "")
                            try:
                                value_cell.value = float(cleaned)
                            except ValueError:
                                pass
                        value_cell.number_format = "_-* #,##0.00_-;_-* -#,##0.00_-;_-* \"-\"??_-;_-@_-"
            elif header_value in integer_columns:
                for cell in worksheet.iter_cols(
                    min_col=col_idx, max_col=col_idx, min_row=2
                ):
                    for value_cell in cell:
                        if isinstance(value_cell.value, str):
                            cleaned = value_cell.value.replace(",", "")
                            try:
                                value_cell.value = float(cleaned)
                            except ValueError:
                                pass
                        value_cell.number_format = "#,##0"
            elif header_value in percent_columns:
                for cell in worksheet.iter_cols(
                    min_col=col_idx, max_col=col_idx, min_row=2
                ):
                    for value_cell in cell:
                        if isinstance(value_cell.value, str):
                            cleaned = value_cell.value.replace("%", "")
                            try:
                                value_cell.value = float(cleaned) / 100
                            except ValueError:
                                pass
                        value_cell.number_format = "0.00%"

        # Conditional formatting with Excel default red/green styles.
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_font = Font(color="006100")
        red_font = Font(color="9C0006")
        max_row = worksheet.max_row
        if max_row >= 2:
            # G: Change (numeric >0 green, <0 red)
            worksheet.conditional_formatting.add(
                f"G2:G{max_row}",
                CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill, font=green_font),
            )
            worksheet.conditional_formatting.add(
                f"G2:G{max_row}",
                CellIsRule(operator="lessThan", formula=["0"], fill=red_fill, font=red_font),
            )

            # H/J/L/P: Up/Down text rules
            worksheet.conditional_formatting.add(
                f"H2:H{max_row}",
                FormulaRule(formula=['$H2="Up"'], fill=green_fill, font=green_font),
            )
            worksheet.conditional_formatting.add(
                f"H2:H{max_row}",
                FormulaRule(formula=['$H2="Down"'], fill=red_fill, font=red_font),
            )

            worksheet.conditional_formatting.add(
                f"J2:J{max_row}",
                FormulaRule(formula=['$J2="Up"'], fill=green_fill, font=green_font),
            )
            worksheet.conditional_formatting.add(
                f"J2:J{max_row}",
                FormulaRule(formula=['$J2="Down"'], fill=red_fill, font=red_font),
            )

            worksheet.conditional_formatting.add(
                f"L2:L{max_row}",
                FormulaRule(formula=['$L2="Up"'], fill=green_fill, font=green_font),
            )
            worksheet.conditional_formatting.add(
                f"L2:L{max_row}",
                FormulaRule(formula=['$L2="Down"'], fill=red_fill, font=red_font),
            )

            worksheet.conditional_formatting.add(
                f"P2:P{max_row}",
                FormulaRule(formula=['$P2="Up"'], fill=green_fill, font=green_font),
            )
            worksheet.conditional_formatting.add(
                f"P2:P{max_row}",
                FormulaRule(formula=['$P2="Down"'], fill=red_fill, font=red_font),
            )


def main() -> None:
    data_dir = Path(__file__).resolve().parent.parent / "data"
    if not data_dir.exists():
        raise FileNotFoundError(f"Data directory not found: {data_dir}")

    files = sorted(data_dir.glob("* Stock Price History.csv"))
    if not files:
        raise FileNotFoundError(
            "No '* Stock Price History.csv' files found in data/."
        )

    for file_path in files:
        print(f"Processing {file_path.name}...")
        process_csv(file_path, data_dir)
    print("Done.")


if __name__ == "__main__":
    main()
