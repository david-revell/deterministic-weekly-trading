"""Generate a latest-week summary Excel file from weekly ETF CSV data."""

from __future__ import annotations

from pathlib import Path

import pandas as pd

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from atr import add_atr_features
from ema_macd import add_ema_macd_features


ATR_BELOW_MULTIPLIER = 0.25
ATR_ABOVE_MULTIPLIER = 1.0
HH_WINDOW = 5


def _coerce_numeric(series: pd.Series) -> pd.Series:
    if series.dtype == object:
        cleaned = series.astype(str).str.replace(",", "", regex=False)
        return pd.to_numeric(cleaned, errors="coerce")
    return series.astype(float)


def _find_column(columns: list[str], names: set[str]) -> str:
    for col in columns:
        if col.lower() in names:
            return col
    raise ValueError(f"Missing required column: {', '.join(sorted(names))}")


def _format_multiplier(value: float) -> str:
    return f"{value:g}"


def _build_summary_tables(
    close_value: float,
    atr_value: float,
    ema5_value: float,
    ema12_value: float,
    ema26_value: float,
    s1_value: float,
    s2_value: float,
    r1_value: float,
    r2_value: float,
    hh_value: float,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    atr_below_label = f"{_format_multiplier(ATR_BELOW_MULTIPLIER)}x ATR (Below)"
    atr_above_label = f"{_format_multiplier(ATR_ABOVE_MULTIPLIER)}x ATR (Above)"

    buy_rows = [
        {
            "Order": "Current",
            "KPI": "Current Close Price",
            "Value": close_value,
            "Chng": 0.0,
        },
        {
            "Order": "Buy",
            "KPI": atr_below_label,
            "Value": close_value - (ATR_BELOW_MULTIPLIER * atr_value),
            "Chng": pd.NA,
        },
        {
            "Order": "Buy",
            "KPI": "0.5x ATR (Below)",
            "Value": close_value - (0.5 * atr_value),
            "Chng": pd.NA,
        },
        {
            "Order": "Buy",
            "KPI": "1x ATR (Below)",
            "Value": close_value - (1.0 * atr_value),
            "Chng": pd.NA,
        },
        {
            "Order": "Buy",
            "KPI": "EMA(5)",
            "Value": ema5_value,
            "Chng": pd.NA,
        },
        {
            "Order": "Buy",
            "KPI": "EMA(12)",
            "Value": ema12_value,
            "Chng": pd.NA,
        },
        {
            "Order": "Buy",
            "KPI": "EMA(26)",
            "Value": ema26_value,
            "Chng": pd.NA,
        },
        {
            "Order": "Buy",
            "KPI": "S1",
            "Value": s1_value,
            "Chng": pd.NA,
        },
        {
            "Order": "Buy",
            "KPI": "S2",
            "Value": s2_value,
            "Chng": pd.NA,
        },
    ]

    sell_rows = [
        {
            "Order": "Current",
            "KPI": "Current Close Price",
            "Value": close_value,
            "Chng": 0.0,
        },
        {
            "Order": "Sell",
            "KPI": atr_above_label,
            "Value": close_value + (ATR_ABOVE_MULTIPLIER * atr_value),
            "Chng": pd.NA,
        },
        {
            "Order": "Sell",
            "KPI": "1.5x ATR (Above)",
            "Value": close_value + (1.5 * atr_value),
            "Chng": pd.NA,
        },
        {
            "Order": "Sell",
            "KPI": "2x ATR (Above)",
            "Value": close_value + (2.0 * atr_value),
            "Chng": pd.NA,
        },
        {
            "Order": "Sell",
            "KPI": "R1",
            "Value": r1_value,
            "Chng": pd.NA,
        },
        {
            "Order": "Sell",
            "KPI": "R2",
            "Value": r2_value,
            "Chng": pd.NA,
        },
        {
            "Order": "Sell",
            "KPI": "HH",
            "Value": hh_value,
            "Chng": pd.NA,
        },
    ]

    current_close = close_value if pd.notna(close_value) else pd.NA
    if pd.notna(current_close):
        for row in buy_rows:
            value = row["Value"]
            if pd.notna(value):
                row["Chng"] = (value - current_close) / current_close
        for row in sell_rows:
            value = row["Value"]
            if pd.notna(value):
                row["Chng"] = (value - current_close) / current_close

    buy_df = pd.DataFrame(buy_rows)
    sell_df = pd.DataFrame(sell_rows)

    buy_df = (
        buy_df.reset_index()
        .sort_values(by="Value", ascending=False, na_position="last")
        .reset_index(drop=True)
        .drop(columns="index")
    )
    sell_df = (
        sell_df.reset_index()
        .sort_values(by="Value", ascending=False, na_position="last")
        .reset_index(drop=True)
        .drop(columns="index")
    )

    return buy_df, sell_df


def process_csv(file_path: Path, output_dir: Path) -> None:
    df = pd.read_csv(file_path)
    if "Price" in df.columns and "Close" not in df.columns:
        df = df.rename(columns={"Price": "Close"})

    close_col = _find_column(list(df.columns), {"close"})
    high_col = _find_column(list(df.columns), {"high"})
    low_col = _find_column(list(df.columns), {"low"})

    df[close_col] = _coerce_numeric(df[close_col])
    df[high_col] = _coerce_numeric(df[high_col])
    df[low_col] = _coerce_numeric(df[low_col])

    # Ensure oldest-first ordering before indicator calculations.
    df = df.iloc[::-1].reset_index(drop=True)
    df = add_ema_macd_features(df, close_column=close_col, spans=(12, 26, 5))
    df = add_atr_features(df, high_column=high_col, low_column=low_col, close_column=close_col)

    pp_series = (df[high_col] + df[low_col] + df[close_col]) / 3
    s1_series = 2 * pp_series - df[high_col]
    s2_series = pp_series - (df[high_col] - df[low_col])
    r1_series = 2 * pp_series - df[low_col]
    r2_series = pp_series + (df[high_col] - df[low_col])
    hh_series = df[high_col].rolling(window=HH_WINDOW, min_periods=HH_WINDOW).max()

    latest = df.iloc[-1]
    buy_df, sell_df = _build_summary_tables(
        close_value=latest[close_col],
        atr_value=latest["atr"],
        ema5_value=latest["ema_5"],
        ema12_value=latest["ema_12"],
        ema26_value=latest["ema_26"],
        s1_value=s1_series.iloc[-1],
        s2_value=s2_series.iloc[-1],
        r1_value=r1_series.iloc[-1],
        r2_value=r2_series.iloc[-1],
        hh_value=hh_series.iloc[-1],
    )

    etf_name = file_path.stem.replace(" Stock Price History", "").replace(" ", "_")
    output_path = output_dir / f"Weekly_{etf_name}_Summary.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl", mode="w") as writer:
        buy_df.to_excel(writer, sheet_name="Summary", index=False, startrow=0, startcol=0)
        sell_df.to_excel(writer, sheet_name="Summary", index=False, startrow=0, startcol=5)
        worksheet = writer.sheets["Summary"]

        base_font = Font(name="Candara Light")
        header_font = Font(name="Candara Light", bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", start_color="16365C", end_color="16365C")
        header_alignment = Alignment(horizontal="left", vertical="center")

        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = base_font

        header_rows = [1]
        header_cols = [1, 6]
        for row in header_rows:
            for start_col in header_cols:
                for col_idx in range(start_col, start_col + 4):
                    cell = worksheet.cell(row=row, column=col_idx)
                    if cell.value is None:
                        continue
                    cell.alignment = header_alignment
                    cell.font = header_font
                    cell.fill = header_fill

        for col_idx in range(1, worksheet.max_column + 1):
            if col_idx == 5:
                continue
            header_value = worksheet.cell(row=1, column=col_idx).value
            header_text = str(header_value) if header_value is not None else ""
            worksheet.column_dimensions[get_column_letter(col_idx)].width = max(
                12, len(header_text) + 2
            )
        worksheet.column_dimensions["E"].width = 3

        value_cols = [3, 8]
        chng_cols = [4, 9]
        for value_col in value_cols:
            for cell in worksheet.iter_cols(
                min_col=value_col, max_col=value_col, min_row=2
            ):
                for value_cell in cell:
                    value_cell.number_format = "#,##0.00"

        for chng_col in chng_cols:
            for cell in worksheet.iter_cols(
                min_col=chng_col, max_col=chng_col, min_row=2
            ):
                for value_cell in cell:
                    value_cell.number_format = "0.00%"

        highlight_fill = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")
        for table_df, start_col in [(buy_df, 0), (sell_df, 5)]:
            if "KPI" not in table_df.columns:
                continue
            current_rows = table_df.index[table_df["KPI"] == "Current Close Price"].tolist()
            for idx in current_rows:
                row_number = 2 + idx
                kpi_col = start_col + 2
                value_col = start_col + 3
                worksheet.cell(row=row_number, column=kpi_col).fill = highlight_fill
                worksheet.cell(row=row_number, column=value_col).fill = highlight_fill


def main() -> None:
    data_dir = Path(__file__).resolve().parent.parent / "data"
    if not data_dir.exists():
        raise FileNotFoundError(f"Data directory not found: {data_dir}")

    files = sorted(data_dir.glob("* Stock Price History.csv"))
    if not files:
        raise FileNotFoundError(
            "No '* Stock Price History.csv' files found in data/."
        )

    for file_path in files:
        print(f"Processing {file_path.name}...")
        process_csv(file_path, data_dir)
    print("Done.")


if __name__ == "__main__":
    main()
